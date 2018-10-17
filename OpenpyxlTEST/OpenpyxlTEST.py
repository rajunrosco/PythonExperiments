# -*- coding: utf8 -*-

# Personal Python 3.6 Template
import getopt
import os
import sys
import shutil
import openpyxl as pyxl
import openpyxl.styles.numbers as Numbers  # the numbers module contains Format definitions


def UnicodeTest1():
    print("Execute TEST")
    if os.path.exists("test.xlm"):
        wb = pyxl.load_workbook("test.xlsm", keep_vba=True)
    else:
        wb = pyxl.Workbook()
    ws = wb.active

    for row in ws.rows:
        for cell in row:
            currentvalue = cell.value
            if(cell.value is not None):
                coordinate = cell.coordinate
                if (cell.column == 'B' ) and (cell.row == 2):
                    cell.value = u'中国 This is pretty cool'
                    print("[%s] %s -> %s" %(coordinate, currentvalue, cell.value))  
                else:
                    print("[%s] %s" %(coordinate, currentvalue))

    wb.save("test.xlsm")

    print("TEST complete...")
    
def UnicodeTest2():
    # test insertion of UTF8 character into a cell
    print("execute TEST2")

    wb = pyxl.Workbook()

    ws = wb.active

    ws['B1'] = '中国'

    wb.save("test2.xlsm")

    print("TEST2 complete...")


def CreateTest3():
    #test format cell to be currency
    
    wb = pyxl.Workbook()
    ws = wb.active

    a4 = ws['A4'] # specify cell
    a4.number_format=Numbers.FORMAT_CURRENCY_USD_SIMPLE # specify cell to be of CURRENCY format usd simple
    a4.value = 222 # specify value for cell

    a5 = ws['A5']
    a5.number_format=Numbers.BUILTIN_FORMATS[7]  # Format accounts for negative currancy surrounded by parenthesis
    a5.value = -222

    wb.save('test3.xlsx')

    print("TEST3 complete...")


from openpyxl import Workbook, load_workbook, worksheet
def LoadTest3():
    wb = load_workbook("test3.xlsx")
    ws = wb.active

    #worksheet.Worksheet.iter_rows()

    last_row = ws.max_row
    last_col = ws.max_column

    xlTableDict = {}
    rownum = 0
    for row in ws.iter_rows(max_row=last_row, max_col=last_col):
        rownum+=1
        rowdata =[]
        for col in row:
            rowdata.append(col.value)
        print("--")
        xlTableDict["row{}".format(rownum)] = tuple(rowdata)

    for key, value in xlTableDict.items():
        print(repr(value))
    print("TEST4 complete...")
    
#Fill a big table	
def CreateBigXL():
	wb = pyxl.Workbook()
	ws = wb.active
	
	for i in range(1,20000):
		ws.cell(row=i, column=1, value="Key"+repr(i))
		ws.cell(row=i, column=2, value="Text"+repr(i))
		if ((i % 100) == 0):
			print( repr(i))
			
	wb.save("testBIG.xlsx")
	
	wb.close()
	
#Read big table	
def ExportBigXL2XML():
	wb = load_workbook("testBIG.xlsx")
	ws = wb.active
	
	f = open('testBIG.xml','w')
	f.write('<Root>\n')
	
	for i in range(1,ws.max_row):
		f.write('	<String>\n')
		for j in range(1,4):
			if j == 1: f.write('		<Key>'+ws.cell(row=i,column=j).value+'</Key>\n')
			if j == 2: f.write('		<Text>'+ws.cell(row=i,column=j).value+'</Text>\n')
			if j == 3:
				#pdb.set_trace()
				currentcell = ws.cell(row=i, column=j)
				if currentcell.value is None:
					f.write('		<Comment></Comment>\n')
		f.write('	</String>\n')
		if ((i % 100) == 0):
			print( repr(i))
			

	f.write('</Root>\n')
	f.close()
	wb.close()

def CellNavigationTest():
    wb = pyxl.Workbook()
    ws = wb.active

    ws["A1"] = "A1"
    temp = ws["A1"].value
    ws.cell(row=1, column=2, value="A2")

    a3cell = ws.cell(row=3, column=1)
    temp = a3cell.coordinate    # returns string coordinate of cell "A3"
    temp = a3cell.column        # returns sting of column "A"



    wb.save("testcell.xlsx")
    wb.close()

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.workbook.defined_name import DefinedName

def RangeTest():
    wb = pyxl.Workbook()
    #ws = wb.active

    ws = wb.create_sheet()
    ws.title = 'Pi'
    ws['F5'] = 3.14
    named_range = wb.create_named_range("Pi",ws, "F5", 0)

    temp = type(named_range)

    wb.defined_names.append("myrange")
    #wb.add_named_range(named_range)


    myrange = wb.defined_names('myrange')
    range_dests = myrange.destinations

    redFill = PatternFill(start_color='FFFF0000',end_color='FFFF0000', fill_type='solid')

    myrange = ws.defined_range(ws["D1":"F10"])
    range_dests = myrange.destinations
    for cell in range_dests:
        cell.fill = redFill

    colC = ws['C']

   
    col_range = ws[2:2]
    for cell in col_range:
        temp = cell.coordinate
        cell.fill = redFill



    row10 = ws[10]
    row_range = ws[5:10]

    wb.save("testrange.xlsx")
    wb.close


######################################################################################################################################################################
#
# Main()
#
######################################################################################################################################################################
    

def Main(argv):

    iStop = 1

    #RangeTest()


# If module is executed by name using python.exe, enter script through Main() method.  If it is imported as a module, Main() is never executed at it is used as a library
if __name__ == "__main__":
    Main(sys.argv)
